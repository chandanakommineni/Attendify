from openpyxl import Workbook, load_workbook
from datetime import datetime

students = ["chandu", "appu"]
file = "attendance.xlsx"


def mark_attendance(recognized):

    date = datetime.now().strftime("%Y-%m")

    try:
        wb = load_workbook(file)
        ws = wb.active
    except:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Date", "Status"])

    # ✅ Read existing rows to avoid duplicates
    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, d, status = row
        existing.add((name, d))

    # ✅ Write only missing attendance rows
    for student in students:
        status = "Present" if student in recognized else "Absent"

        if (student, date) not in existing:
            ws.append([student, date, status])

    wb.save(file)
